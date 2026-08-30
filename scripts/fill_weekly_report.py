#!/usr/bin/env python3
"""填写实习生周报 xlsx 模板。

推荐用法（内容来自技能第四步生成的周报 .md，Windows cmd 下无引号转义问题）:
  python fill_weekly_report.py \
    --template assets/实习生周报标准模板.xlsx \
    --from-md "输出目录/姓名工作周报(YYYYMMDD-YYYYMMDD)-指导老师.md" \
    --sheet "9月第1周" \
    --output "输出目录/姓名工作周报(YYYYMMDD-YYYYMMDD)-指导老师.xlsx"

备用用法（命令行直接传 JSON，bash 引号写法；cmd 下转义麻烦，不推荐）:
  python fill_weekly_report.py \
    --template assets/实习生周报标准模板.xlsx \
    --output 输出路径.xlsx \
    --name "姓名" --department "部门" --position "岗位" \
    --mentor "指导老师" --entry-date "2026-08-04" --fill-date "2026-08-14" \
    --sheet "8月第3周" \
    --tasks '[["内容1","完成情况1"],["内容2","完成情况2"]]' \
    --learnings '["收获1","收获2"]' \
    --problems '["问题1"]' \
    --plans '["计划1","计划2","计划3"]'

--from-md 与命令行参数可混用：显式给出的命令行参数覆盖 md 中解析到的值
（--sheet 不在 md 中，始终必传）。

条目行数完全由内容驱动：
- 内容多于模板预留行 → 在板块末尾插入新行（复制上一行的样式与合并单元格）
- 内容少于模板预留行 → 删除板块末尾多余的空白行
- 行高随内容自适应：内容超过两行时抬高行高，避免文字被截断
"""

import argparse
import json
import math
import re
import shutil
import sys
from copy import copy
from datetime import datetime
from pathlib import Path


def _setup_stdio():
    """Windows 下管道输出默认跟随 GBK 代码页，打印中文/emoji 会崩，主动切到 UTF-8。"""
    for stream in (sys.stdout, sys.stderr):
        if stream is not None and hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8", errors="replace")


_setup_stdio()

# 尝试导入 openpyxl，若不可用则给出安装提示
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 需要 openpyxl 库。请运行: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ── 板块表头识别文本（按模板中出现的顺序）──────────────
SECTION_HEADERS = [
    ("tasks", "一、本周工作内容"),
    ("learn", "二、学习与收获"),
    ("problem", "三、问题与困难"),
    ("plan", "四、下周工作计划"),
]

# md 信息区可出现的字段（顺序即输出顺序）
MD_FIELDS = ("姓名", "部门", "岗位", "指导老师", "入职时间", "填写日期")

# 允许的日期输入格式，解析成功后统一写入为 yyyy/m/d
DATE_FORMATS = ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日", "%Y%m%d")

# 行高标定：模板条目行为微软雅黑 10pt，两行内容行高 30 → 每行 15
LINE_HEIGHT = 15.0
DEFAULT_COL_WIDTH = 8.43  # openpyxl 中未显式设宽的列按 Excel 默认宽度计


def main():
    args = _parse_args()

    if args.from_md:
        _merge_md(args, parse_md(args.from_md))

    if not args.name:
        sys.exit("错误: 缺少姓名。请用 --from-md 提供周报 md 文件，或用 --name 指定。")

    # 解析并校验 JSON 数据
    tasks     = _parse_json_list(args.tasks, "--tasks", pair=True)
    learnings = _parse_json_list(args.learnings, "--learnings")
    problems  = _parse_json_list(args.problems, "--problems")
    plans     = _parse_json_list(args.plans, "--plans")
    for label, data in (("工作内容", tasks), ("学习与收获", learnings),
                        ("问题与困难", problems), ("下周计划", plans)):
        if not data:
            print(f"⚠️ 板块「{label}」内容为空，该板块将不保留条目行")

    # 复制模板 → 输出（输出目录不存在时自动创建）
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(args.template, out_path)
    wb = openpyxl.load_workbook(out_path)

    # 定位或创建分表
    ws = _get_or_create_sheet(wb, args.sheet)

    # 填写基本信息
    _fill_basic_info(ws, args)

    # 动态定位四个板块的数据区
    sections = _locate_sections(ws)

    # 从后往前调整行数并填写（每个板块的调整只影响其自身数据区之后的行，
    # 因此后处理的板块行号不受影响）
    _resize_and_fill(ws, sections["plan"], plans)
    _resize_and_fill(ws, sections["problem"], problems)
    _resize_and_fill(ws, sections["learn"], learnings)
    _resize_and_fill(ws, sections["tasks"], tasks, kind="tasks")

    wb.save(out_path)
    print(f"✅ 周报已生成: {out_path}")
    print(f"   姓名={args.name} | 分表={args.sheet} | 任务 {len(tasks)} 条、"
          f"收获 {len(learnings)} 条、问题 {len(problems)} 条、计划 {len(plans)} 条")


# ── 参数解析 ──────────────────────────────────────────

def _parse_args():
    p = argparse.ArgumentParser(description="填写实习生周报 xlsx 模板")
    p.add_argument("--template", required=True, help="模板 xlsx 路径")
    p.add_argument("--output", required=True, help="输出 xlsx 路径（目录不存在会自动创建）")
    p.add_argument("--from-md", dest="from_md", default=None,
                   help="周报 .md 文件路径（本技能第四步定义的格式），"
                        "自动解析全部基本信息与内容；显式给出的命令行参数优先")
    p.add_argument("--name", default="", help="姓名 → B2（可由 --from-md 提供）")
    p.add_argument("--department", default="", help="部门（末级组织）→ D2")
    p.add_argument("--position", default="", help="岗位 → F2")
    p.add_argument("--mentor", default="", help="指导老师 → B3")
    p.add_argument("--entry-date", dest="entry_date", default="",
                   help="入职时间 → D3（支持 2026-08-04 / 2026/8/4 / 2026年8月4日 等）")
    p.add_argument("--fill-date", dest="fill_date", default="",
                   help="填写日期 → F3（支持格式同上）")
    p.add_argument("--sheet", required=True, help='分表名称，如 "9月第1周"')
    p.add_argument("--tasks", default=None,
                   help='JSON 数组，每项为 [工作内容, 完成情况]')
    p.add_argument("--learnings", default=None,
                   help='JSON 字符串数组，不含序号前缀')
    p.add_argument("--problems", default=None,
                   help='JSON 字符串数组，不含序号前缀')
    p.add_argument("--plans", default=None,
                   help='JSON 字符串数组，不含序号前缀')
    return p.parse_args()


def _merge_md(args, md):
    """--from-md 解析结果作为默认值，命令行显式参数覆盖。"""
    for arg_key in ("name", "department", "position", "mentor",
                    "entry_date", "fill_date"):
        if not getattr(args, arg_key):
            setattr(args, arg_key, md[arg_key])
    for arg_key in ("tasks", "learnings", "problems", "plans"):
        if getattr(args, arg_key) is None:
            setattr(args, arg_key, json.dumps(md[arg_key], ensure_ascii=False))


def _parse_json_list(raw, arg_name, pair=False):
    """解析 JSON 数组参数并做友好校验。pair=True 时每项须为二元字符串数组。"""
    if raw is None:
        return []
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        sys.exit(f"错误: {arg_name} 不是合法 JSON（{e}）。"
                 f"Windows cmd 下建议改用 --from-md 传 md 文件。")
    if not isinstance(data, list):
        sys.exit(f"错误: {arg_name} 应为 JSON 数组，实际为: {data!r}")
    for i, item in enumerate(data, 1):
        if pair:
            if not (isinstance(item, list) and len(item) == 2
                    and all(isinstance(x, str) for x in item)):
                sys.exit(f"错误: {arg_name} 第 {i} 项应为 [\"工作内容\", \"完成情况\"] "
                         f"二元数组，实际为: {item!r}")
        elif not isinstance(item, str):
            sys.exit(f"错误: {arg_name} 第 {i} 项应为字符串，实际为: {item!r}")
    return data


# ── 周报 .md 解析 ─────────────────────────────────────

# 信息区字段：值惰性匹配，遇到"空白 + 下一个字段名"或行尾结束，
# 兼容模板的 4 空格分隔与单空格分隔两种写法
_INFO_FIELD_RE = re.compile(
    r"(姓名|部门|岗位|指导老师|入职时间|填写日期)：\s*(.*?)\s*"
    r"(?=\s+(?:姓名|部门|岗位|指导老师|入职时间|填写日期)：|$)"
)


def parse_md(md_path):
    """解析技能第四步模板生成的周报 .md，返回与命令行参数同构的字典。"""
    text = Path(md_path).read_text(encoding="utf-8")
    lines = text.splitlines()

    # 信息区：引用块（> 开头）中的 "字段：值"
    info = {f: "" for f in MD_FIELDS}
    for line in lines:
        if not line.lstrip().startswith(">"):
            continue
        s = line.lstrip().lstrip(">").strip()
        for m in _INFO_FIELD_RE.finditer(s):
            info[m.group(1)] = m.group(2).strip()

    # 四个板块：以 "## 一、…" 二级标题切分
    sections = {}
    current = None
    for line in lines:
        m = re.match(r"^##(?!#)\s*(.*)$", line)
        if m:
            current = None
            for key, htext in SECTION_HEADERS:
                if htext in m.group(1):
                    current = key
                    sections[current] = []
                    break
        elif current is not None:
            sections[current].append(line)

    missing = [t for k, t in SECTION_HEADERS if k not in sections]
    if missing:
        sys.exit(f"错误: md 中未找到板块标题 {missing}，"
                 f"请确认 md 由本技能第四步的模板生成: {md_path}")
    if not info["姓名"]:
        sys.exit(f"错误: md 信息区未解析到姓名（应存在 '> 姓名：…' 行）: {md_path}")

    parsed = {
        "name": info["姓名"],
        "department": info["部门"],
        "position": info["岗位"],
        "mentor": info["指导老师"],
        "entry_date": info["入职时间"],
        "fill_date": info["填写日期"],
        "tasks": _md_table_rows(sections["tasks"]),
        "learnings": _md_numbered_items(sections["learn"], "二、学习与收获"),
        "problems": _md_numbered_items(sections["problem"], "三、问题与困难"),
        "plans": _md_numbered_items(sections["plan"], "四、下周工作计划"),
    }
    return parsed


def _md_table_rows(rows):
    """解析"一、本周工作内容"的表格体，返回 [[工作内容, 完成情况], …]。"""
    items = []
    for line in rows:
        s = line.strip()
        if not s.startswith("|"):
            continue
        cells = [c.strip().replace("\\|", "|") for c in s.strip("|").split("|")]
        if len(cells) < 3:
            continue
        if cells[0] == "序号" or set(cells[0]) <= set(":-"):  # 列头行/分隔行
            continue
        content, completion = "|".join(cells[1:-1]), cells[-1]
        if content or completion:
            items.append([content, completion])
    return items


def _md_numbered_items(rows, label):
    """解析编号列表板块（兼容 "1、" 与 "1. " 两种编号写法）。"""
    items = []
    for line in rows:
        m = re.match(r"^\s*\d+\s*[、.]\s*(.+?)\s*$", line)
        if m:
            items.append(m.group(1))
    if not items and any(line.strip() for line in rows):
        print(f"⚠️ md 板块「{label}」有内容但未解析到编号条目（应使用 1、2、… 编号）")
    return items


# ── 分表管理 ──────────────────────────────────────────

def _get_or_create_sheet(wb, sheet_name):
    """返回目标分表；若不存在则从干净模板复制新建。"""
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    # 选择最干净的 sheet 作为复制源（优先选无预填数据的）
    source_sheet = _pick_cleanest_sheet(wb)
    new_ws = wb.copy_worksheet(source_sheet)
    new_ws.title = sheet_name

    # 清理复制源可能带过来的示例占位数据
    _clean_example_data(new_ws)

    return new_ws


def _pick_cleanest_sheet(wb):
    """在模板的所有分表中选数据最'干净'的作为复制源。
    判断标准：B2（姓名）为空或为示例文本 且 B3（指导老师）为空或为示例。
    """
    best = wb[wb.sheetnames[0]]
    for name in wb.sheetnames:
        ws = wb[name]
        b2 = str(ws['B2'].value or '')
        b3 = str(ws['B3'].value or '')
        # 完全空白的最优先
        if b2 == '' and b3 == '':
            return ws
        # 含"示例"的次之
        if '示例' in b2 or '示例' in b3:
            best = ws
    return best


def _clean_example_data(ws):
    """清除复制源可能残留的示例占位文本（基本信息区）。

    列表板块中的 "1、" "2、" 占位符无需在此清理：填写时会覆盖，
    多余行会被删除（见 _resize_and_fill）。
    """
    clean_cells = ['B2', 'D2', 'F2', 'B3', 'D3', 'F3']
    for ref in clean_cells:
        val = str(ws[ref].value or '')
        if '示例' in val:
            ws[ref].value = None


# ── 板块定位 ──────────────────────────────────────────

def _locate_sections(ws):
    """扫描 A 列，按表头文本动态定位四个板块的数据区。

    返回 {key: {"start": row, "end": row}}，数据区不含表头行本身：
    - 前三个板块以相邻板块的表头行为边界
    - 任务板块额外跳过列头行（"序号/工作内容/完成情况"）
    - 最后一个板块以"说明"文字行（或 max_row）为边界
    """
    header_rows = {}
    for row in range(1, ws.max_row + 1):
        val = str(ws.cell(row=row, column=1).value or '')
        for key, text in SECTION_HEADERS:
            if text in val and key not in header_rows:
                header_rows[key] = row

    missing = [text for key, text in SECTION_HEADERS if key not in header_rows]
    if missing:
        sys.exit(f"错误: 模板中未找到板块表头 {missing}，请确认模板结构")

    # 最后一个板块之后的边界：说明文字起始行
    last_header = header_rows[SECTION_HEADERS[-1][0]]
    tail = ws.max_row + 1
    for row in range(last_header + 1, ws.max_row + 1):
        val = str(ws.cell(row=row, column=1).value or '')
        if val and '说明' in val:
            tail = row
            break

    sections = {}
    for i, (key, _) in enumerate(SECTION_HEADERS):
        start = header_rows[key] + 1
        if i + 1 < len(SECTION_HEADERS):
            end = header_rows[SECTION_HEADERS[i + 1][0]] - 1
        else:
            end = tail - 1
        if key == "tasks":
            # 跳过"序号/工作内容/完成情况"列头行
            col_header = str(ws.cell(row=start, column=1).value or '')
            if '序号' in col_header:
                start += 1
        sections[key] = {"start": start, "end": end}
    return sections


# ── 行数调整 ──────────────────────────────────────────

def _resize_and_fill(ws, section, items, kind="list"):
    """把板块数据区调整为恰好 len(items) 行并填写内容。

    - 内容多于预留行 → 在数据区末尾插入新行（复制样式与合并单元格）
    - 内容少于预留行 → 删除数据区末尾多余的空白行
    - 填写后按内容抬高行高（只增不减），避免长文本被固定行高截断
    """
    start, end = section["start"], section["end"]
    needed = len(items)
    current = end - start + 1

    if current < 1:
        sys.exit(f"错误: 板块数据区为空（第 {start} 行起），请确认模板结构")

    if needed > current:
        _insert_rows(ws, end + 1, needed - current)
    elif needed < current:
        _delete_rows(ws, start + needed, current - needed)

    # 填写内容（数据区起始行不变，始终写入前 needed 行）
    for i in range(needed):
        row = start + i
        if kind == "tasks":
            content, completion = items[i]
            ws.cell(row=row, column=1, value=i + 1)      # A: 序号
            ws.cell(row=row, column=2, value=content)    # B: 工作内容
            ws.cell(row=row, column=4, value=completion) # D: 完成情况
            _fit_row_height(ws, row, [
                (content, _merged_width(ws, row, 2)),      # B:C 合并
                (completion, _merged_width(ws, row, 4)),   # D:F 合并
            ])
        else:
            text = f"{i + 1}、{items[i]}"
            ws.cell(row=row, column=1, value=text)         # A:F 合并
            _fit_row_height(ws, row, [(text, _merged_width(ws, row, 1))])


# ── 行高自适应 ────────────────────────────────────────

def _column_width(ws, col_idx):
    # 模板中相邻等宽列会合并为一个范围型 ColumnDimension（如 B~C 共用一个宽度），
    # 必须按 min~max 范围匹配；直接按列字母取值会拿到自动创建的默认宽度
    for dim in ws.column_dimensions.values():
        if dim.width and dim.min is not None and dim.max is not None \
                and dim.min <= col_idx <= dim.max:
            return dim.width
    dim = ws.column_dimensions.get(get_column_letter(col_idx))
    return dim.width if dim is not None and dim.width else DEFAULT_COL_WIDTH


def _merged_width(ws, row, col):
    """(row, col) 所在合并区域的总列宽；未合并则为该列列宽。"""
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return sum(_column_width(ws, c) for c in range(rng.min_col, rng.max_col + 1))
    return _column_width(ws, col)


def _fit_row_height(ws, row, texts_widths):
    """按内容估算所需行数并抬高行高（只增不减）。

    列宽单位约等于一个半角字符，中文等全角字符按 2 个单位计；
    预留 2 个单位作单元格内边距。
    """
    max_lines = 1
    for text, width in texts_widths:
        if not text:
            continue
        capacity = max(width - 2.0, 4.0)
        lines = 0
        for seg in str(text).split("\n"):
            units = sum(2.0 if ord(ch) > 127 else 1.0 for ch in seg)
            lines += max(1, math.ceil(units / capacity))
        max_lines = max(max_lines, lines)
    needed = max_lines * LINE_HEIGHT
    dim = ws.row_dimensions[row]
    if dim.height is None or needed > dim.height:
        dim.height = needed


def _shift_row_dimensions(ws, first_row, last_row, delta):
    """openpyxl 的 insert_rows/delete_rows 不会平移 row_dimensions，手动维护。

    将 [first_row, last_row] 区间外、首行及之后的行高整体平移 delta
    （delta>0 为插入场景的行号后移，delta<0 为删除场景的行号前移）。
    """
    heights = {r: d.height for r, d in ws.row_dimensions.items()}
    for r in [r for r in ws.row_dimensions if r >= first_row]:
        del ws.row_dimensions[r]
    for r, h in heights.items():
        if h is None:
            continue
        if r < first_row:
            ws.row_dimensions[r].height = h
        elif delta > 0 or r > last_row:
            # 插入：首行及之后全部后移；删除：仅删除区之后的行前移
            ws.row_dimensions[r + delta].height = h


def _insert_rows(ws, row_idx, count):
    """在 row_idx 处插入 count 行。

    1. 将 row_idx 及之后的合并区域整体下移（跨插入点的区域扩大）
    2. 新行复制 row_idx-1 行的样式、行高与单行合并区域
    3. row_dimensions 随行号平移
    """
    template_row = row_idx - 1

    # 收集全部合并区域（clear 后需要完整重建）
    unchanged = []  # 插入点之前，保持不动 (min_col, min_row, max_col, max_row)
    shifted = []    # 插入点及之后，整体下移
    copied = []     # 上一行的单行合并列范围 (min_col, max_col)
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = rng.bounds
        if min_row >= row_idx:
            shifted.append((min_col, min_row + count, max_col, max_row + count))
        elif max_row >= row_idx:
            # 合并区域跨插入点 → 向下扩大
            shifted.append((min_col, min_row, max_col, max_row + count))
        else:
            unchanged.append((min_col, min_row, max_col, max_row))
        if min_row == template_row and max_row == template_row:
            copied.append((min_col, max_col))

    ws.merged_cells.ranges.clear()
    _shift_row_dimensions(ws, row_idx, row_idx - 1, count)
    ws.insert_rows(row_idx, count)
    for min_col, min_row, max_col, max_row in unchanged + shifted:
        ws.merge_cells(start_row=min_row, start_column=min_col,
                       end_row=max_row, end_column=max_col)

    # 新行复制样式、行高与合并
    for i in range(count):
        r = row_idx + i
        for col in range(1, ws.max_column + 1):
            dst = ws.cell(row=r, column=col)
            dst._style = copy(ws.cell(row=template_row, column=col)._style)
        if ws.row_dimensions[template_row].height:
            ws.row_dimensions[r].height = ws.row_dimensions[template_row].height
        for min_col, max_col in copied:
            ws.merge_cells(start_row=r, start_column=min_col,
                           end_row=r, end_column=max_col)


def _delete_rows(ws, first_row, count):
    """删除 first_row 起的 count 行，并同步维护合并区域与行高。

    模板中所有合并区域均为单行合并，被删的空白条目行对应整行区域，
    因此只处理"完全在删除区内（丢弃）"与"完全在删除区后（上移）"两种情况。
    """
    last_row = first_row + count - 1
    kept = []
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = rng.bounds
        if min_row >= first_row and max_row <= last_row:
            continue  # 完全位于删除区内 → 丢弃
        if min_row > last_row:
            min_row -= count
            max_row -= count
        elif max_row >= first_row:
            # 起点在删除区上方、终点落入删除区 → 截断到删除区上方
            max_row = first_row - 1
            if max_row < min_row:
                continue
        kept.append((min_col, min_row, max_col, max_row))

    ws.merged_cells.ranges.clear()
    _shift_row_dimensions(ws, first_row, last_row, -count)
    ws.delete_rows(first_row, count)
    for min_col, min_row, max_col, max_row in kept:
        ws.merge_cells(start_row=min_row, start_column=min_col,
                       end_row=max_row, end_column=max_col)


# ── 基本信息 ──────────────────────────────────────────

def _fill_basic_info(ws, args):
    """填写姓名 / 部门 / 岗位 / 指导老师 / 入职时间 / 填写日期。"""
    ws['B2'] = args.name
    ws['D2'] = args.department
    ws['F2'] = args.position
    ws['B3'] = args.mentor

    # 日期字段：尝试解析为 datetime，保留 yyyy/m/d 格式
    _set_date_cell(ws, 'D3', args.entry_date)
    _set_date_cell(ws, 'F3', args.fill_date)


def _set_date_cell(ws, cell_ref, date_str):
    """将单元格设为日期值（datetime）并应用 yyyy/m/d 数字格式。

    支持多种常见写法；均无法解析时原样写入字符串。
    """
    if not date_str:
        return
    date_str = str(date_str).strip()
    for fmt in DATE_FORMATS:
        try:
            dt = datetime.strptime(date_str, fmt)
            ws[cell_ref] = dt
            ws[cell_ref].number_format = 'yyyy/m/d'
            return
        except ValueError:
            continue
    ws[cell_ref] = date_str  # 无法解析则原样写入


if __name__ == "__main__":
    main()
